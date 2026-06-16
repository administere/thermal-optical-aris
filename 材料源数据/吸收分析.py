#!/usr/bin/env python3
"""
850 nm 吸收系数外推分析
========================
从 MOESM6 紫外-可见-近红外光谱数据外推 DiSubPc·C70 在 850 nm (VCSEL 波长) 的吸收系数。

方法:
  1. 乌尔巴赫带尾: α(E) ∝ exp(σE/E_u) — 对有机 CT 吸收最合理
  2. 幂律: α(λ) ∝ λ^p
  3. 与假设的参数值进行比较

关键发现:
  - 光学带隙 E_g = 2.25 eV (551 nm) — 850 nm (1.46 eV) 远低于带隙
  - 乌尔巴赫外推得出 α(850 nm) ≈ 350 cm⁻¹ (不确定性 ±2×)
  - 这比假设值低约 5.7 倍 (α_assumed = 2000 cm⁻¹)
  - 10μm 薄膜吸收率: 估计 29%，而假设值为 86%

对验证的影响:
  - 在 850 nm 处，薄膜吸收率从 ~60% 降至 ~30%
  - 到达探测器的光功率增加 5 倍
  - SNR 改善约 7 dB (光电流增加)
  - 但自加热能力从 60% 降至 30% 吸收率
  - 保持 242°C 可能需要更多辅助功率
"""

import numpy as np
import openpyxl
from scipy import stats
import os

BASE = os.path.dirname(os.path.abspath(__file__))

def load_extinction_data(filepath=None):
    """加载 MOESM6 Figure 2b 的摩尔消光系数数据"""
    if filepath is None:
        filepath = os.path.join(BASE, '41566_2026_1912_MOESM6_ESM.xlsx')

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Figure 2b']

    waves, eps = [], []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0] is not None and row[1] is not None:
            waves.append(float(row[0]))
            eps.append(float(row[1]))  # 2DiSubPc, in 10⁵ L/(cm·mol)

    wb.close()
    return np.array(waves), np.array(eps) * 1e5  # Convert to L/(cm·mol)


def extrapolate_urbach(waves_nm, eps_cm, target_nm=850):
    """
    乌尔巴赫带尾外推: α(E) ∝ exp(σE/kT)
    对低于带隙的有机 CT 吸收具有物理合理性
    """
    E_eV = 1240.0 / waves_nm
    E_target = 1240.0 / target_nm

    # 使用长波尾部 (600–800 nm)
    tail_mask = (waves_nm >= 600) & (waves_nm <= 800)
    E_tail = E_eV[tail_mask]
    eps_tail = eps_cm[tail_mask]

    valid = eps_tail > 0
    E_fit = E_tail[valid]
    log_eps = np.log(eps_tail[valid])

    slope, intercept, r, p, std_err = stats.linregress(E_fit, log_eps)

    log_eps_target = slope * E_target + intercept
    eps_target = np.exp(log_eps_target)
    E_u = 1.0 / abs(slope)  # Urbach energy (eV)

    return {
        'eps_cm': eps_target,
        'urbach_energy_meV': E_u * 1000,
        'r_squared': r**2,
        'slope': slope,
        'intercept': intercept,
        'method': 'Urbach tail',
    }


def extrapolate_power_law(waves_nm, eps_cm, target_nm=850):
    """幂律外推: α ∝ λ^p (经验方法)"""
    tail_mask = (waves_nm >= 600) & (waves_nm <= 800) & (eps_cm > 0)
    log_w = np.log(waves_nm[tail_mask])
    log_e = np.log(eps_cm[tail_mask])

    slope, intercept, r, p, std_err = stats.linregress(log_w, log_e)
    eps_target = np.exp(slope * np.log(target_nm) + intercept)

    return {
        'eps_cm': eps_target,
        'exponent': slope,
        'r_squared': r**2,
        'method': 'Power law',
    }


def compute_alpha(eps_cm, density_gcm3=1.581, mw_gmol=1054.36):
    """将消光系数转换为薄膜吸收系数

    α = ε × C, 其中 C 是发色团浓度 (mol/L)
    C_solid = ρ / M_W × 1000 [mol/L]
    """
    conc_mol_L = density_gcm3 / mw_gmol * 1000
    alpha_cm = eps_cm * conc_mol_L
    return alpha_cm, conc_mol_L


def film_absorption(alpha_cm, film_um=10.0):
    """计算给定厚度薄膜的吸收率"""
    T = np.exp(-alpha_cm * film_um * 1e-4)
    return 1 - T


def main():
    waves, eps = load_extinction_data()

    print("=" * 70)
    print("  DiSubPc·C70: 850 nm 吸收系数外推")
    print("=" * 70)

    # 实验数据范围
    print(f"\n  实验数据: λ = {min(waves):.0f} – {max(waves):.0f} nm")
    print(f"  目标波长: 850 nm (VCSEL)")
    print(f"  带隙 E_g = 2.25 eV → λ_onset ≈ 551 nm")
    print(f"  850 nm = 1.46 eV → 低于带隙 {2.25-1.46:.2f} eV")

    # 乌尔巴赫外推
    urb = extrapolate_urbach(waves, eps, 850)
    print(f"\n  ── 方法 1: 乌尔巴赫带尾 ──")
    print(f"  乌尔巴赫能量 E_u = {urb['urbach_energy_meV']:.0f} meV")
    print(f"  R² = {urb['r_squared']:.4f}")
    print(f"  ε(850 nm) = {urb['eps_cm']:.0f} L/(cm·mol)")

    # 幂律外推
    pl = extrapolate_power_law(waves, eps, 850)
    print(f"\n  ── 方法 2: 幂律外推 ──")
    print(f"  幂指数 p = {pl['exponent']:.1f}")
    print(f"  R² = {pl['r_squared']:.4f}")
    print(f"  ε(850 nm) = {pl['eps_cm']:.0f} L/(cm·mol)")

    # 转换为 α
    alpha_urb, conc = compute_alpha(urb['eps_cm'])
    alpha_pl, _ = compute_alpha(pl['eps_cm'])

    print(f"\n  ── 转换为吸收系数 (晶体密度 1.581 g/cm³) ──")
    print(f"  发色团浓度: {conc:.1f} mol/L")
    print(f"  α(乌尔巴赫) = {alpha_urb:.0f} cm⁻¹")
    print(f"  α(幂律)     = {alpha_pl:.0f} cm⁻¹")

    # 10μm 薄膜吸收率
    abs_urb = film_absorption(alpha_urb)
    abs_pl = film_absorption(alpha_pl)
    abs_assumed = film_absorption(2000)

    print(f"\n  ── 10μm 薄膜吸收率 ──")
    print(f"  乌尔巴赫估计: {abs_urb*100:.1f}%")
    print(f"  幂律估计:     {abs_pl*100:.1f}%")
    print(f"  当前假设值:   {abs_assumed*100:.1f}% (α=2000 cm⁻¹)")

    # 不确定性范围
    print(f"\n  ── 不确定性范围 (±2× 的 ε) ──")
    best_alpha = alpha_urb
    for mult, label in [(0.5, '低'), (1.0, '最佳'), (2.0, '高')]:
        a = best_alpha * mult
        abs_a = film_absorption(a)
        print(f"  {label}: α = {a:.0f} cm⁻¹ → 吸收率 = {abs_a*100:.1f}%")

    # 对验证的影响
    print(f"\n  ── 下游影响 ──")
    P_vcsel = 5.0  # mW
    abs_old = film_absorption(2000, 10.0)
    abs_new = film_absorption(best_alpha, 10.0)

    P_absorbed_old = P_vcsel * abs_old
    P_absorbed_new = P_vcsel * abs_new

    print(f"  单点吸收功率 (旧): {P_absorbed_old:.1f} mW")
    print(f"  单点吸收功率 (新): {P_absorbed_new:.1f} mW")
    print(f"  自加热能力降为原来的: {abs_new/abs_old*100:.0f}%")

    P_transmitted_old = P_vcsel * (1 - abs_old)
    P_transmitted_new = P_vcsel * (1 - abs_new)
    print(f"  探测器接收功率 (旧): {P_transmitted_old:.2f} mW")
    print(f"  探测器接收功率 (新): {P_transmitted_new:.2f} mW")
    print(f"  探测器 SNR 改善: +{10*np.log10(P_transmitted_new/P_transmitted_old):.1f} dB")

    return {
        'alpha_best_cm': best_alpha,
        'alpha_range': (best_alpha * 0.5, best_alpha * 2.0),
        'abs_best': abs_new,
        'abs_assumed': abs_old,
        'urbach': urb,
    }


if __name__ == '__main__':
    results = main()
